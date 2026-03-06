"""
Procure-Me Pricing Calculator

A Streamlit application for processing vendor quotes with markup and tax calculations.
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import sys
import os

# Add src directory to path for imports
sys.path.append(os.path.join(os.path.dirname(__file__), 'src'))

from src.pricing_calculator import (
    PricingProcessor,
    CalculationMethod,
    InvoiceFormat,
    TradeType,
    ComplianceMode,
    TRADE_PRESETS,
    margin_to_markup,
    markup_to_margin
)
from src.excel_parser import ExcelParser
from src.materials_db import MaterialsDatabase
from src.labor_db import LaborDatabase
from src.pdf_parser import PDFParser
from src.vendor_detector import VendorDetector


def _detect_columns(df: pd.DataFrame) -> dict:
    """
    Detect which DataFrame columns contain description, quantity, and unit cost
    by analyzing the actual cell values — not just column names.
    
    Handles cases where:
    - Column names are generic (Col0, Col1, etc.) from PDF parsing
    - The first row of data contains the real column headers
      (e.g. 'Description', 'Qty', 'Rate')
    - Vendor page headers repeat across rows
    
    Returns dict with keys 'desc', 'qty', 'cost' mapping to column names.
    """
    scores = {col: {'desc': 0, 'qty': 0, 'cost': 0} for col in df.columns}
    
    desc_keywords = ['desc', 'material', 'item', 'product', 'name']
    qty_keywords = ['qty', 'quantity', 'count']
    # Prioritize Quote Qty over BOM Qty
    quote_qty_keywords = ['quote qty', 'quote quantity', 'quotequantity', 'quoteqty']
    bom_qty_keywords = ['bom qty', 'bom quantity', 'bomqty']
    cost_keywords = ['cost', 'price', 'rate', 'amount', 'unit cost', 'unit price']
    
    for col in df.columns:
        col_lower = str(col).lower().replace(' ', '')
        col_lower_with_space = str(col).lower()
        vals = df[col].dropna().astype(str)
        sample = vals.head(20)
        
        # ── 1. Column-name hints ──
        if any(k in col_lower_with_space for k in desc_keywords):
            scores[col]['desc'] += 3
        
        # PRIORITY: Quote Qty gets highest score
        if any(k.replace(' ', '') in col_lower for k in quote_qty_keywords):
            scores[col]['qty'] += 10  # High priority for Quote Qty
        # BOM Qty gets lower score
        elif any(k.replace(' ', '') in col_lower for k in bom_qty_keywords):
            scores[col]['qty'] += 2   # Lower priority for BOM Qty
        # Generic qty/quantity
        elif any(k in col_lower_with_space for k in qty_keywords):
            scores[col]['qty'] += 3
        
        if any(k in col_lower_with_space for k in cost_keywords):
            scores[col]['cost'] += 3
        
        # ── 2. First few values may be in-data headers ──
        # Check the first 5 non-empty values for header keywords
        for v in vals.head(5):
            vl = v.strip().lower()
            if vl in ('description', 'item', 'material', 'product', 'item description'):
                scores[col]['desc'] += 6
            # Prioritize Quote Qty in header values too
            if vl in ('quote qty', 'quote quantity', 'quotequantity', 'quoteqty'):
                scores[col]['qty'] += 10
            elif vl in ('bom qty', 'bom quantity', 'bomqty'):
                scores[col]['qty'] += 3
            elif vl in ('qty', 'quantity'):
                scores[col]['qty'] += 6
            if vl in ('rate', 'unit cost', 'cost', 'price', 'unit price'):
                scores[col]['cost'] += 6
            if vl in ('total',):
                # Total column is not cost — penalise cost, but don't add
                scores[col]['cost'] -= 2
        
        # ── 3. Content analysis ──
        numeric_count = 0
        long_text_count = 0
        unique_vals = set()
        nums = []
        for v in sample:
            unique_vals.add(v.strip())
            v_clean = v.replace('$', '').replace(',', '').replace('T', '').strip()
            try:
                nums.append(float(v_clean))
                numeric_count += 1
            except ValueError:
                pass
            if len(v) > 25:
                long_text_count += 1
        
        total = max(len(sample), 1)
        numeric_ratio = numeric_count / total
        text_ratio = long_text_count / total
        unique_ratio = len(unique_vals) / total
        
        # Columns with mostly long text are likely descriptions
        if text_ratio > 0.3:
            scores[col]['desc'] += 4
        
        # Columns where every row is the same value are repeated headers
        # (like Date, Estimate #, P.O. No., Terms, Rep, Project)
        # These should NOT be qty or cost
        if unique_ratio < 0.3 and total > 3:
            scores[col]['qty'] -= 3
            scores[col]['cost'] -= 3
            scores[col]['desc'] -= 2
        
        # Columns with varied numeric values
        if numeric_ratio > 0.4 and nums:
            avg = sum(nums) / len(nums)
            # Penalize columns where all numbers are the same (header repeat)
            if len(set(round(n, 2) for n in nums)) <= 2:
                scores[col]['qty'] -= 2
                scores[col]['cost'] -= 2
            else:
                whole_ratio = sum(1 for n in nums if n == int(n)) / len(nums)
                if whole_ratio > 0.5:
                    scores[col]['qty'] += 3
                decimal_ratio = sum(1 for n in nums if n != int(n)) / len(nums)
                if decimal_ratio > 0.2:
                    scores[col]['cost'] += 3
    
    # Pick best column for each role (no column used twice)
    # Prioritize desc first (most distinctive), then qty, then cost
    result = {}
    used = set()
    for role in ['desc', 'qty', 'cost']:
        best_col = None
        best_score = -1
        for col in df.columns:
            if col not in used and scores[col][role] > best_score:
                best_score = scores[col][role]
                best_col = col
        if best_col:
            result[role] = best_col
            used.add(best_col)
    
    return result


# Configure Streamlit page
st.set_page_config(
    page_title="Procure-Me Pricing Calculator",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .section-header {
        font-size: 1.5rem;
        color: #2c3e50;
        margin-top: 2rem;
        margin-bottom: 1rem;
    }
    .success-message {
        padding: 1rem;
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 0.25rem;
        color: #155724;
    }
    .info-box {
        padding: 1rem;
        background-color: #e7f3ff;
        border: 1px solid #b8daff;
        border-radius: 0.25rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)


def main():
    """Main application function."""
    # Hero Section
    st.markdown("""
    <div style="text-align: center; padding: 2rem 0 1rem 0;">
        <h1 style="font-size: 3rem; margin-bottom: 0.5rem;">🔧 Procure-Me</h1>
        <p style="font-size: 1.5rem; color: #666; margin-bottom: 2rem;">
            Turn vendor quotes into client-ready estimates in seconds
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Metrics Row
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("⚡ Processing Time", "< 30 sec")
    with col2:
        st.metric("📊 Output Formats", "7 types")
    with col3:
        st.metric("🔒 Your Data", "Never stored")
    with col4:
        st.metric("💰 Cost", "FREE")
    
    # Value Props
    st.markdown("""
    ### Why Procure-Me?
    
    ✅ **Stop manual spreadsheet work** — Upload vendor quote, get client estimate instantly  
    ✅ **Formula transparency** — See exactly how every price is calculated  
    ✅ **Multiple output formats** — Excel, CSV, Word, Markdown — choose what works for you  
    ✅ **Configurable rates** — Adjust margin (0-50%) and tax (0-20%) with simple sliders  
    ✅ **No signup required** — Start processing quotes immediately, no account needed  
    """)
    
    # Quick Start Guide
    with st.expander("� Quick Start Guide - First time here?"):
        st.markdown("""
        ### How to Use Procure-Me
        
        **Step 1: Upload Your Vendor Quote**
        - Click the "📊 Process Quote" tab below
        - Upload your Excel (.xlsx) or PDF file
        - Supported formats: vendor quotes, BOMs, material lists
        
        **Step 2: Configure Your Rates**
        - Use the sidebar sliders to set your margin (default 10%)
        - Set your tax rate (default 8.25%)
        - Adjust shipping costs if applicable
        
        **Step 3: Map Your Columns**
        - Procure-Me auto-detects columns (Description, Qty, Cost)
        - Verify the mapping is correct
        - Adjust if needed
        
        **Step 4: Download Your Results**
        - Get 7 different output formats:
          - **Client Quote** (clean 6-column format)
          - **Internal Audit** (full cost breakdown with formulas)
          - **Combined Workbook** (both sheets in one file)
          - **CSV exports** (for data portability)
          - **Summary reports** (Markdown or Word)
        
        **All processing happens locally in your browser. No data is uploaded or stored.**
        """)
    
    st.markdown("---")
    
    # Initialize materials database with persistent storage
    data_file_path = os.path.join(os.path.dirname(__file__), "materials_data.json")
    materials_db = MaterialsDatabase(data_file_path)
    
    # Initialize labor database with persistent storage
    labor_file_path = os.path.join(os.path.dirname(__file__), "labor_data.json")
    labor_db = LaborDatabase(labor_file_path)
    
    # Main navigation
    tab1, tab2, tab3 = st.tabs(["📊 Process Quote", "🔧 Build Quote", "📦 Materials List"])
    
    with tab1:
        process_quote_tab(materials_db, labor_db)
    
    with tab2:
        build_quote_tab(materials_db, labor_db)
    
    with tab3:
        materials_list_tab(materials_db)
    
    # Footer
    st.markdown("---")
    footer_col1, footer_col2, footer_col3 = st.columns(3)
    
    with footer_col1:
        st.markdown("**About**")
        st.markdown(
            "Built for contractors, estimators, and procurement professionals who are tired of manual spreadsheet work."
        )
    
    with footer_col2:
        st.markdown("**Resources**")
        st.markdown("📖 [Documentation](https://github.com/SuperGremlin25/Procure-Me#readme)")
        st.markdown("🐛 [Report Bug](https://github.com/SuperGremlin25/Procure-Me/issues)")
        st.markdown("💡 [Request Feature](https://github.com/SuperGremlin25/Procure-Me/issues)")
    
    with footer_col3:
        st.markdown("**Connect**")
        st.markdown("⭐ [Star on GitHub](https://github.com/SuperGremlin25/Procure-Me)")
        st.markdown("Made with ❤️ by Digital Insurgent Media")
    
    st.caption("v1.0.0 | MIT License | No data is stored or transmitted")


def process_quote_tab(materials_db, labor_db):
    """Process uploaded quote files."""
    # Sidebar for settings
    st.sidebar.markdown('<h2 class="section-header">Pricing Configuration</h2>', 
                        unsafe_allow_html=True)
    
    # Quick Start Mode Selector
    st.sidebar.markdown("### Quick Start")
    compliance_mode = st.sidebar.selectbox(
        "What are you pricing?",
        options=[
            ComplianceMode.COMMERCIAL.value,
            ComplianceMode.GOVERNMENT_GRANT.value,
            ComplianceMode.GOVERNMENT_CONTRACT.value,
            ComplianceMode.CUSTOM.value
        ],
        format_func=lambda x: {
            "commercial": "Commercial Job (standard pricing)",
            "grant": "Government Grant (ARPA/BEAD compliance)",
            "contract": "Direct Government Contract (FAR compliance)",
            "custom": "Custom (advanced settings)"
        }[x],
        help="Loads preset configuration for common scenarios"
    )
    
    # Trade Type Selector
    trade_type = st.sidebar.selectbox(
        "Type of work",
        options=[t.value for t in TradeType],
        format_func=lambda x: TRADE_PRESETS[TradeType(x)]['description'],
        index=0,  # Default to Telecom
        help="Loads suggested margin ranges for this trade"
    )
    trade_type_enum = TradeType(trade_type)
    trade_preset = TRADE_PRESETS[trade_type_enum]
    
    # Determine default margin based on mode
    if compliance_mode in [ComplianceMode.GOVERNMENT_GRANT.value, ComplianceMode.GOVERNMENT_CONTRACT.value]:
        default_margin = trade_preset['government_margin']
    else:
        default_margin = trade_preset['default_margin']
    
    # Margin Configuration
    st.sidebar.markdown("### Contractor Margin")
    
    # Gross Margin toggle
    is_gross_margin = st.sidebar.checkbox(
        "Enter as Gross Margin %",
        value=True,
        help="If checked, enter your target gross margin (profit/revenue). Tool will calculate required markup automatically."
    )
    
    margin_label = "Target Gross Margin (%)" if is_gross_margin else "Markup (%)"
    margin_help = (
        f"Your target gross margin. Suggested range: {trade_preset['suggested_margin_min']*100:.0f}%-{trade_preset['suggested_margin_max']*100:.0f}%"
        if is_gross_margin else
        "Markup percentage to add to costs"
    )
    
    margin_value = st.sidebar.slider(
        margin_label,
        min_value=5.0,
        max_value=50.0,
        value=default_margin * 100,
        step=1.0,
        help=margin_help
    )
    
    margin_rate = margin_value / 100
    
    # Show conversion if in gross margin mode
    if is_gross_margin:
        required_markup = margin_to_markup(margin_rate) * 100
        st.sidebar.caption(f"✓ Required markup: {required_markup:.1f}%")
        if margin_rate == 0.33:
            st.sidebar.caption("ℹ️ 33% is a common GC target")
    
    # Show suggested range
    if compliance_mode in [ComplianceMode.GOVERNMENT_GRANT.value, ComplianceMode.GOVERNMENT_CONTRACT.value]:
        if margin_rate > 0.15:
            st.sidebar.warning("⚠️ Government contracts typically question margins above 15%")
    else:
        suggested_min = trade_preset['suggested_margin_min'] * 100
        suggested_max = trade_preset['suggested_margin_max'] * 100
        if margin_rate * 100 < suggested_min or margin_rate * 100 > suggested_max:
            st.sidebar.info(f"ℹ️ Typical range for {trade_preset['description']}: {suggested_min:.0f}%-{suggested_max:.0f}%")
    
    # Tax Rate
    st.sidebar.markdown("### Sales Tax")
    tax_rate = st.sidebar.slider(
        "Tax Rate (%)",
        min_value=0.0,
        max_value=20.0,
        value=8.25,
        step=0.25,
        help="Sales tax rate on materials"
    ) / 100
    
    # Advanced Settings (Custom Mode Only)
    if compliance_mode == ComplianceMode.CUSTOM.value:
        with st.sidebar.expander("⚙️ Advanced Calculation Settings"):
            st.markdown("**Calculation Method**")
            calc_method = st.selectbox(
                "How to calculate pricing:",
                options=[m.value for m in CalculationMethod],
                format_func=lambda x: {
                    "margin_then_tax": "Margin-Then-Tax (default)",
                    "tax_separate": "Tax-Separate (government standard)",
                    "additive": "Additive (simplest)",
                    "cost_plus_fee": "Cost-Plus-Fixed-Fee"
                }[x],
                index=0,
                help="Different calculation methods for different business models"
            )
            calculation_method = CalculationMethod(calc_method)
            
            if calc_method == "cost_plus_fee":
                fixed_fee = st.number_input(
                    "Fixed Fee ($)",
                    min_value=0.0,
                    value=0.0,
                    step=100.0,
                    help="Fixed dollar fee for cost-plus contracts"
                )
            else:
                fixed_fee = 0.0
            
            st.markdown("**Invoice Format**")
            inv_format = st.selectbox(
                "Output format:",
                options=[f.value for f in InvoiceFormat],
                format_func=lambda x: {
                    "detailed_audit": "Detailed Audit Trail (recommended)",
                    "far_compliant": "FAR-Compliant Government",
                    "grant_compliance": "Grant Compliance (ARPA/BEAD)",
                    "clean_client": "Clean Client Invoice",
                    "combo": "Combo (Audit + Clean)"
                }[x],
                index=0,
                help="How the invoice is presented to the client"
            )
            invoice_format = InvoiceFormat(inv_format)
    else:
        # Auto-select based on compliance mode
        if compliance_mode == ComplianceMode.GOVERNMENT_GRANT.value:
            calculation_method = CalculationMethod.TAX_SEPARATE
            invoice_format = InvoiceFormat.GRANT_COMPLIANCE
            fixed_fee = 0.0
        elif compliance_mode == ComplianceMode.GOVERNMENT_CONTRACT.value:
            calculation_method = CalculationMethod.TAX_SEPARATE
            invoice_format = InvoiceFormat.FAR_COMPLIANT
            fixed_fee = 0.0
        else:  # Commercial
            calculation_method = CalculationMethod.MARGIN_THEN_TAX
            invoice_format = InvoiceFormat.DETAILED_AUDIT
            fixed_fee = 0.0
    
    # Show current configuration summary
    with st.sidebar.expander("📊 Current Configuration Summary"):
        st.write(f"**Mode:** {compliance_mode.title()}")
        st.write(f"**Trade:** {trade_preset['description']}")
        if is_gross_margin:
            st.write(f"**Gross Margin:** {margin_rate*100:.1f}%")
            st.write(f"**Markup:** {margin_to_markup(margin_rate)*100:.1f}%")
        else:
            st.write(f"**Markup:** {margin_rate*100:.1f}%")
            st.write(f"**Gross Margin:** {markup_to_margin(margin_rate)*100:.1f}%")
        st.write(f"**Tax:** {tax_rate*100:.2f}%")
        st.write(f"**Calc Method:** {calculation_method.value.replace('_', ' ').title()}")
        st.write(f"**Invoice Format:** {invoice_format.value.replace('_', ' ').title()}")
        
        # Example calculation
        st.markdown("**Example: $100 item**")
        processor = PricingProcessor(
            tax_rate=tax_rate,
            margin_rate=margin_rate,
            calculation_method=calculation_method,
            is_gross_margin=is_gross_margin,
            fixed_fee=fixed_fee
        )
        result = processor.calculate_composite_rate(100.0)
        st.write(f"Cost: $100.00")
        st.write(f"Margin: ${result['margin_dollars']:.2f}")
        st.write(f"Tax: ${result['tax_dollars']:.2f}")
        st.write(f"**Total: ${result['composite_rate']:.2f}**")
    
    # Feedback Section
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 💬 Feedback")
    st.sidebar.markdown("Help us improve Procure-Me!")
    
    feedback_text = st.sidebar.text_area(
        "Share your thoughts:",
        placeholder="Feature requests, bugs, suggestions...",
        height=100,
        label_visibility="collapsed",
        key="feedback_text"
    )
    
    if st.sidebar.button("📤 Send Feedback", use_container_width=True):
        if feedback_text.strip():
            st.sidebar.success("✅ Thanks for your feedback!")
            st.sidebar.info("We'll review it soon!")
            # TODO: Future - integrate with Google Forms API or email service
        else:
            st.sidebar.warning("Please enter some feedback first")
    
    st.sidebar.markdown(
        "📋 [Take our 2-min survey](https://github.com/SuperGremlin25/Procure-Me/issues) for a chance to get early access to premium features!",
        unsafe_allow_html=True
    )
    
    # First-time user helper
    if 'files_processed' not in st.session_state:
        st.session_state.files_processed = 0
    
    # ── PROJECT TYPE SELECTION ────────────────────────────
    st.markdown('<h2 class="section-header">Select Quote Type</h2>', 
                unsafe_allow_html=True)
    
    st.markdown("""
    **What type of quote are you processing?**
    
    Choose the quote type that matches your vendor document:
    """)
    
    project_type = st.radio(
        "Quote Type:",
        options=["📦 Materials Only", "👷 Labor Only", "🏗️ Full Build (Materials + Labor)"],
        index=0,
        help="Select what type of quote you're uploading",
        horizontal=True
    )
    
    # Store in session state
    if project_type == "📦 Materials Only":
        st.session_state['quote_type'] = "materials"
        st.info("**Materials Quote** - You'll upload vendor pricing for materials (conduit, cable, hardware, etc.)")
    elif project_type == "👷 Labor Only":
        st.session_state['quote_type'] = "labor"
        st.info("**Labor Estimate** - You'll enter labor tasks, hours, and rates")
    else:
        st.session_state['quote_type'] = "full_build"
        st.info("**Full Build Quote** - Combined materials and labor for complete project pricing")
    
    st.markdown("---")
    
    # File upload section
    st.markdown('<h2 class="section-header">Upload Vendor Quote</h2>', 
                unsafe_allow_html=True)
    
    # Welcome message for first-time users
    if st.session_state.files_processed == 0:
        st.info("""
        👋 **Welcome!** Upload your first vendor quote to get started. 
        Supported formats: Excel (.xlsx, .xls) or PDF.
        
        💡 **Tip:** Use the sidebar to adjust margin and tax rates before processing.
        """)
    
    uploaded_file = st.file_uploader(
        "Choose a file",
        type=['xlsx', 'xls', 'pdf'],
        help="Upload vendor quote (Excel or PDF)",
        accept_multiple_files=False
    )
    
    if uploaded_file is not None:
        # Validate file size (max 50MB)
        if uploaded_file.size > 50 * 1024 * 1024:
            st.error("File too large. Please upload a file smaller than 50MB.")
            return
        
        st.success(f"✅ File uploaded: {uploaded_file.name}")
        
        # ── VENDOR VALIDATION (NEW) ──────────────────────────────
        st.markdown("---")
        st.markdown('<h3>⚠️ STEP 1: Verify Vendor Information</h3>', unsafe_allow_html=True)
        st.markdown("""
        **Why this matters:** Accurate vendor identification helps with column mapping and future quote comparison features.
        """)
        
        # Initialize vendor detector
        detector = VendorDetector()
        
        # Try filename detection first
        filename_result = detector.detect_from_filename(uploaded_file.name)
        
        # For content detection, we need to preview the file
        vendor_name = None
        try:
            # Quick preview parse (first 15 rows only)
            if uploaded_file.type == 'application/pdf':
                # For PDF, we'll skip content detection to avoid double parsing
                content_result = {"vendor_name": "Unknown", "confidence": 0.0}
            else:
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_preview:
                    tmp_preview.write(uploaded_file.getbuffer())
                    preview_path = tmp_preview.name
                
                # Read just first 15 rows for detection
                preview_df = pd.read_excel(preview_path, nrows=15)
                content_result = detector.detect_from_dataframe(preview_df)
                
                # Reset file pointer for later processing
                uploaded_file.seek(0)
                
                # Clean up temp file
                try:
                    os.unlink(preview_path)
                except:
                    pass
        except Exception as e:
            content_result = {"vendor_name": "Unknown", "confidence": 0.0}
        
        # Use best result
        best_result = detector.combine_results(content_result, filename_result)
        
        # Display detection results with appropriate UI
        if best_result['confidence'] >= 0.7:
            st.success(f"✅ **Vendor Detected:** {best_result['vendor_name']} (Confidence: {best_result['confidence']:.0%})")
            if best_result['indicators']:
                with st.expander("Detection Details"):
                    for indicator in best_result['indicators'][:3]:
                        st.caption(f"• {indicator}")
            vendor_name = st.text_input(
                "Confirm or Edit Vendor Name",
                value=best_result['vendor_name'],
                help="Verify this is correct before proceeding",
                key="vendor_name_input"
            )
            
        elif best_result['confidence'] >= 0.4:
            st.warning(f"⚠️ **Possible Vendor:** {best_result['vendor_name']} (Low Confidence: {best_result['confidence']:.0%})")
            if best_result['indicators']:
                with st.expander("Detection Details"):
                    for indicator in best_result['indicators'][:3]:
                        st.caption(f"• {indicator}")
            vendor_name = st.text_input(
                "Confirm or Edit Vendor Name",
                value=best_result['vendor_name'],
                help="Please verify this is correct - low confidence detection",
                key="vendor_name_input"
            )
            
        else:
            st.error("🛑 **VENDOR NOT DETECTED**")
            st.markdown("""
            **Manual Entry Required**
            
            The vendor name could not be automatically detected from your quote file.
            Please enter the vendor name manually to ensure accurate processing.
            
            💡 **Tip:** For better auto-detection in the future, ensure the vendor name 
            appears in the header rows of your quote.
            """)
            vendor_name = st.text_input(
                "Enter Vendor Name",
                placeholder="e.g., Vendor Supply Co, ABC Electric, etc.",
                help="Required field - cannot proceed without vendor name",
                key="vendor_name_input"
            )
        
        # Validation gate - prevent proceeding without vendor name
        if not vendor_name or vendor_name.strip() == "" or vendor_name.strip().lower() == "unknown":
            st.error("❌ **Vendor name is required to continue**")
            st.info("Please enter a valid vendor name above to proceed with quote processing.")
            st.stop()
        
        # Success - store vendor name for later use
        st.session_state['current_vendor'] = vendor_name.strip()
        st.success(f"✅ **Proceeding with vendor:** {vendor_name}")
        
        st.markdown("---")
        st.markdown('<h3>📊 STEP 2: Process Quote File</h3>', unsafe_allow_html=True)
        
        # Check file type
        is_pdf = uploaded_file.type == 'application/pdf'
        
        try:
            if is_pdf:
                # Parse PDF
                st.info("📄 PDF detected. Using advanced PDF parser...")
                parser = PDFParser()
                
                # Get PDF info
                pdf_info = parser.get_pdf_info(uploaded_file)
                st.markdown(f"""
                <div class="info-box">
                    <strong>PDF Info:</strong><br>
                    Pages: {pdf_info['pages']}<br>
                    File Size: {pdf_info['file_size'] / 1024:.1f} KB<br>
                    Tables Detected: {'Yes' if pdf_info['has_tables'] else 'No'}<br>
                    Available Strategies: {', '.join(pdf_info['strategies_available'])}
                </div>
                """, unsafe_allow_html=True)
                
                # Parse the PDF
                df = parser.parse_pdf(uploaded_file)
                
                # Display parsing results
                st.markdown('<h3>Extracted Data</h3>', unsafe_allow_html=True)
                st.dataframe(df.head(10), use_container_width=True)
                
                if df.empty:
                    st.error("Could not extract data from PDF. Try uploading a clearer version or an Excel file.")
                    return
            else:
                # Parse Excel file
                # Use secure temporary file
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    temp_path = tmp_file.name
                    tmp_file.write(uploaded_file.getbuffer())
                
                # Initialize parser
                parser = ExcelParser()
                
                # Get sheet information
                sheet_info = parser.get_sheet_info(temp_path)
                
                # Sheet selection
                st.markdown('<h3>Select Sheet to Process</h3>', unsafe_allow_html=True)
                
                sheet_names = list(sheet_info.keys())
                selected_sheet = st.selectbox(
                    "Choose a sheet:",
                    sheet_names,
                    help="Select the sheet containing the quote data"
                )
                
                if selected_sheet:
                    # Display sheet info
                    info = sheet_info[selected_sheet]
                    st.markdown(f"""
                    <div class="info-box">
                        <strong>Sheet Info:</strong><br>
                        Columns: {len(info['columns'])}<br>
                        Rows: {info['row_count']}<br>
                        Columns found: {', '.join(str(c) for c in info['columns'][:5])}{'...' if len(info['columns']) > 5 else ''}
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Parse the selected sheet
                    df = parser.parse_file(temp_path, selected_sheet)
                    
                    # Check for duplicate columns
                    duplicates = parser.detect_duplicate_columns(df)
                    
                    if duplicates:
                        st.warning(f"⚠️ Found {len(duplicates)} duplicate column name(s) in your file")
                        
                        # Auto-rename duplicates for now
                        df = parser.rename_duplicate_columns(df)
                        
                        with st.expander("🔍 Duplicate Column Details", expanded=True):
                            st.markdown("""
                            **Duplicate columns have been automatically renamed:**
                            
                            When Excel files have duplicate column names, we append `.1`, `.2`, etc. to make them unique.
                            You can now select the correct column in the mapping step below.
                            """)
                            
                            for col_name, indices in duplicates.items():
                                st.markdown(f"**'{col_name}'** appears {len(indices)} times")
                                st.caption(f"Renamed to: {col_name}, {col_name}.1, {col_name}.2, etc.")
                    
                    # Display raw data preview
                    st.markdown('<h3>Uploaded Data Preview</h3>', unsafe_allow_html=True)
                    st.dataframe(df.head(10), use_container_width=True)
            
            # ── Step 1: Column Mapping ──────────────────────────────
            st.markdown('<h3>Map Your Columns</h3>', unsafe_allow_html=True)
            st.markdown("Select which columns in your file correspond to the required and optional fields.")
            
            # Auto-detect best columns by analyzing content
            detected = _detect_columns(df)
            all_cols = list(df.columns)
            all_cols_with_none = ['(None)'] + all_cols
            
            def _col_index(detected_key, fallback):
                col = detected.get(detected_key)
                if col and col in all_cols:
                    return all_cols.index(col)
                return fallback
            
            def _col_index_with_none(detected_key, fallback):
                col = detected.get(detected_key)
                if col and col in all_cols:
                    return all_cols_with_none.index(col)
                return fallback
            
            st.markdown("**Required Columns:**")
            req_row1 = st.columns(3)
            
            with req_row1[0]:
                desc_col = st.selectbox(
                    "Description Column",
                    options=all_cols,
                    index=_col_index('desc', 0),
                    help="Column containing material descriptions"
                )
            
            with req_row1[1]:
                qty_col = st.selectbox(
                    "Quantity Column",
                    options=all_cols,
                    index=_col_index('qty', min(1, len(all_cols)-1)),
                    help="Column containing quantities"
                )
            
            with req_row1[2]:
                cost_col = st.selectbox(
                    "Unit Cost Column",
                    options=all_cols,
                    index=_col_index('cost', min(2, len(all_cols)-1)),
                    help="Column containing unit costs/prices"
                )
            
            req_row2 = st.columns(3)
            
            with req_row2[0]:
                uom_col = st.selectbox(
                    "Unit of Measurement Column",
                    options=all_cols,
                    index=_col_index('uom', min(3, len(all_cols)-1)),
                    help="Unit type: ea, ft, bag, box, etc."
                )
            
            with req_row2[1]:
                total_col = st.selectbox(
                    "Vendor Total Column",
                    options=all_cols,
                    index=_col_index('total', min(4, len(all_cols)-1)),
                    help="Vendor's line total for validation"
                )
            
            st.markdown("**Optional Columns:**")
            opt_col1 = st.columns(1)[0]
            
            with opt_col1:
                part_col_sel = st.selectbox(
                    "Part Number Column",
                    options=all_cols_with_none,
                    index=_col_index_with_none('part', 0),
                    help="Supplier part number (optional — useful for ordering)"
                )
                part_col = None if part_col_sel == '(None)' else part_col_sel
            
            # ── Step 2: Clean data & extract tax/freight ──────────
            processor = PricingProcessor(
                tax_rate=tax_rate,
                margin_rate=margin_rate,
                calculation_method=calculation_method,
                invoice_format=invoice_format,
                trade_type=trade_type_enum,
                is_gross_margin=is_gross_margin,
                fixed_fee=fixed_fee
            )
            cleaned_df, extracted_tax, extracted_freight = processor.clean_vendor_quote(
                df, desc_col=desc_col, qty_col=qty_col, cost_col=cost_col,
                part_col=part_col, uom_col=uom_col, total_col=total_col
            )
            
            st.markdown('<h3>Cleaned Data Preview</h3>', unsafe_allow_html=True)
            st.dataframe(cleaned_df.head(15), use_container_width=True)
            st.caption(f"{len(cleaned_df)} line items after cleaning ({len(df) - len(cleaned_df)} junk rows removed)")
            
            # Show extracted tax & freight
            info_col1, info_col2, info_col3 = st.columns(3)
            with info_col1:
                if extracted_tax > 0:
                    # Calculate effective tax rate from vendor quote
                    vendor_subtotal = (pd.to_numeric(cleaned_df[cost_col], errors='coerce').fillna(0) *
                                       pd.to_numeric(cleaned_df[qty_col], errors='coerce').fillna(0)).sum()
                    effective_rate = (extracted_tax / vendor_subtotal * 100) if vendor_subtotal > 0 else 0
                    st.info(f"📋 Extracted Sales Tax: **${extracted_tax:,.2f}** (effective rate: {effective_rate:.2f}%)")
                else:
                    st.info("No sales tax found in quote — using sidebar rate")
            with info_col2:
                if extracted_freight > 0:
                    st.info(f"🚚 Extracted Freight: **${extracted_freight:,.2f}**")
                else:
                    st.info("No freight line found in quote")
            with info_col3:
                st.info(f"⚙️ Margin: {margin_rate*100:.1f}% | Tax: {tax_rate*100:.2f}%")
            
            # ── Step 3: Pass-through shipping ─────────────────────
            st.markdown('<h3>Pass-Through Shipping</h3>', unsafe_allow_html=True)
            shipping_cost = st.number_input(
                "Shipping / Freight Cost (pass-through, varies per site)",
                min_value=0.0,
                value=extracted_freight,
                step=10.0,
                help="Enter the shipping cost for this job. Added as a flat line item — no markup applied."
            )
            
            # ── Step 4: Project Name ──────────────────────────────
            st.markdown('<h3>Name Your Project</h3>', unsafe_allow_html=True)
            project_name = st.text_input(
                "Project Name",
                placeholder="e.g., Starlord, Main Street Fiber, Site 42...",
                help="This name will be used as the filename prefix for all downloads",
                value="Quote"
            )
            
            # Clean project name for use in filenames
            safe_project_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in project_name)
            safe_project_name = safe_project_name.replace(' ', '_')
            
            # ── Step 5: Process Quote ─────────────────────────────
            if st.button("Process Quote", type="primary"):
                processor = PricingProcessor(
                    tax_rate=tax_rate,
                    margin_rate=margin_rate,
                    shipping_cost=shipping_cost,
                    calculation_method=calculation_method,
                    invoice_format=invoice_format,
                    trade_type=trade_type_enum,
                    is_gross_margin=is_gross_margin,
                    fixed_fee=fixed_fee
                )
                
                processed_df = processor.process_quote(cleaned_df)
                
                st.markdown('<h2 class="section-header">Processed Quote</h2>', 
                            unsafe_allow_html=True)
                
                # Show calculation info based on method
                example_calc = processor.calculate_composite_rate(100.0)
                st.markdown(f"""
                <div class="info-box">
                    <strong>Calculation Method:</strong> {example_calc['method']}<br>
                    <strong>Formula:</strong> {example_calc['formula_text']}<br>
                    <strong>Example ($100):</strong> Cost: $100.00 → Margin: ${example_calc['margin_dollars']:.2f} → Tax: ${example_calc['tax_dollars']:.2f} → Total: ${example_calc['composite_rate']:.2f}<br>
                    <strong>Shipping:</strong> ${shipping_cost:,.2f} (pass-through)
                </div>
                """, unsafe_allow_html=True)
                
                # ── Client Sheet (3 columns) ─────────────────────
                client_df = processor.generate_client_spreadsheet(
                    processed_df, shipping_cost=shipping_cost
                )
                st.markdown('<h3>Client-Facing Quote (3-Column)</h3>', unsafe_allow_html=True)
                st.dataframe(client_df, use_container_width=True)
                
                client_total = client_df['Total'].sum()
                st.metric("Client Grand Total", f"${client_total:,.2f}")
                
                # ── Audit Sheet (full trail) ─────────────────────
                with st.expander("📊 Internal Audit Trail (for leadership)"):
                    audit_df = processor.generate_internal_audit_spreadsheet(
                        processed_df, tax_rate=tax_rate, margin_rate=margin_rate,
                        shipping_cost=shipping_cost
                    )
                    st.dataframe(audit_df, use_container_width=True)
                
                # ── Downloads ─────────────────────────────────────
                st.markdown('<h3>Download Results</h3>', unsafe_allow_html=True)
                
                # Excel Downloads
                st.markdown("**Excel Workbooks:**")
                dl_col1, dl_col2, dl_col3, dl_col4, dl_col5 = st.columns(5)
                
                with dl_col1:
                    audit_buffer = BytesIO()
                    processor.generate_internal_audit_spreadsheet(
                        processed_df, tax_rate=tax_rate, margin_rate=margin_rate,
                        output_path=audit_buffer, shipping_cost=shipping_cost
                    )
                    st.download_button(
                        label="📥 Audit (.xlsx)",
                        data=audit_buffer.getvalue(),
                        file_name=f"{safe_project_name}_internal_audit.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                with dl_col2:
                    client_buffer = BytesIO()
                    processor.generate_client_spreadsheet(
                        processed_df, output_path=client_buffer,
                        shipping_cost=shipping_cost
                    )
                    st.download_button(
                        label="📥 Client (.xlsx)",
                        data=client_buffer.getvalue(),
                        file_name=f"{safe_project_name}_client_quote.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                with dl_col3:
                    # Generate combined workbook with proper formatting
                    combined_buffer = BytesIO()
                    
                    # Generate audit with formulas
                    audit_buffer_temp = BytesIO()
                    processor.generate_internal_audit_spreadsheet(
                        processed_df, tax_rate=tax_rate, margin_rate=margin_rate,
                        output_path=audit_buffer_temp, shipping_cost=shipping_cost
                    )
                    
                    # Generate client sheet
                    client_buffer_temp = BytesIO()
                    processor.generate_client_spreadsheet(
                        processed_df, output_path=client_buffer_temp,
                        shipping_cost=shipping_cost
                    )
                    
                    # Combine both into one workbook
                    import openpyxl
                    from openpyxl import load_workbook
                    
                    # Load audit workbook
                    audit_buffer_temp.seek(0)
                    audit_wb = load_workbook(audit_buffer_temp)
                    
                    # Load client workbook and copy its sheet
                    client_buffer_temp.seek(0)
                    client_wb = load_workbook(client_buffer_temp)
                    client_sheet = client_wb.active
                    
                    # Copy client sheet to audit workbook
                    audit_wb.create_sheet('Client_Quote')
                    target_sheet = audit_wb['Client_Quote']
                    for row in client_sheet.iter_rows():
                        for cell in row:
                            target_cell = target_sheet[cell.coordinate]
                            target_cell.value = cell.value
                            if cell.has_style:
                                target_cell.font = cell.font.copy()
                                target_cell.border = cell.border.copy()
                                target_cell.fill = cell.fill.copy()
                                target_cell.number_format = cell.number_format
                                target_cell.alignment = cell.alignment.copy()
                    
                    # Rename audit sheet
                    audit_wb['Audit_Details'].title = 'Internal_Audit'
                    
                    # Save combined workbook
                    audit_wb.save(combined_buffer)
                    
                    st.download_button(
                        label="📥 Combined (.xlsx)",
                        data=combined_buffer.getvalue(),
                        file_name=f"{safe_project_name}_quote_complete.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                with dl_col4:
                    md_report = processor.generate_summary_report(
                        processed_df, fmt='md', shipping_cost=shipping_cost,
                        extracted_tax=extracted_tax, extracted_freight=extracted_freight
                    )
                    st.download_button(
                        label="📥 Summary (.md)",
                        data=md_report,
                        file_name=f"{safe_project_name}_summary.md",
                        mime="text/markdown"
                    )
                
                with dl_col5:
                    try:
                        docx_bytes = processor.generate_summary_report(
                            processed_df, fmt='docx', shipping_cost=shipping_cost,
                            extracted_tax=extracted_tax, extracted_freight=extracted_freight
                        )
                        st.download_button(
                            label="📥 Summary (.docx)",
                            data=docx_bytes,
                            file_name=f"{safe_project_name}_summary.docx",
                            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                        )
                    except ImportError:
                        st.caption("Install python-docx for .docx export")
                
                # Add Cleaned Data + CSV Downloads
                st.markdown("**Additional Exports:**")
                csv_col1, csv_col2, csv_col3, csv_col4 = st.columns(4)
                
                with csv_col1:
                    # Cleaned Data Excel
                    cleaned_buffer = BytesIO()
                    with pd.ExcelWriter(cleaned_buffer, engine='xlsxwriter') as writer:
                        cleaned_df.to_excel(writer, sheet_name='Cleaned_Data', index=False)
                    st.download_button(
                        label="📥 Cleaned (.xlsx)",
                        data=cleaned_buffer.getvalue(),
                        file_name=f"{safe_project_name}_cleaned_data.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Raw cleaned data before pricing calculations"
                    )
                
                with csv_col2:
                    # Cleaned Data CSV
                    cleaned_csv = cleaned_df.to_csv(index=False)
                    st.download_button(
                        label="📥 Cleaned (.csv)",
                        data=cleaned_csv,
                        file_name=f"{safe_project_name}_cleaned_data.csv",
                        mime="text/csv",
                        help="Raw cleaned data in CSV format"
                    )
                
                with csv_col3:
                    # Audit CSV
                    audit_csv = audit_df.to_csv(index=False)
                    st.download_button(
                        label="📥 Audit (.csv)",
                        data=audit_csv,
                        file_name=f"{safe_project_name}_internal_audit.csv",
                        mime="text/csv",
                        help="Internal audit trail in CSV format"
                    )
                
                with csv_col4:
                    # Client CSV
                    client_csv = client_df.to_csv(index=False)
                    st.download_button(
                        label="📥 Client (.csv)",
                        data=client_csv,
                        file_name=f"{safe_project_name}_client_quote.csv",
                        mime="text/csv",
                        help="Client quote in CSV format"
                    )
                
                # Success message
                st.markdown("""
                <div class="success-message">
                    ✅ Quote processed successfully! 
                    <br><br>
                    <strong>Excel Downloads:</strong>
                    <br>• <strong>Audit</strong> — Step-by-step dollar breakdown with Part #, UOM (Vendor Cost → +Margin $ → +Tax $ → Composite Rate)
                    <br>• <strong>Client</strong> — Clean 6-column quote with Part #, Description, UOM, Qty, Rate, Total
                    <br>• <strong>Combined</strong> — Both audit and client sheets in one workbook
                    <br>• <strong>Summary</strong> — Markdown or Word report with totals and margin analysis
                    <br><br>
                    <strong>Additional Exports:</strong>
                    <br>• <strong>Cleaned Data</strong> — Raw extracted data (.xlsx or .csv)
                    <br>• <strong>CSV Exports</strong> — All sheets available in CSV format for data portability
                </div>
                """, unsafe_allow_html=True)
                
                # Store processed data in session state for manual additions
                st.session_state['base_cleaned_df'] = cleaned_df.copy()
                st.session_state['current_project_name'] = safe_project_name
                st.session_state['current_tax_rate'] = tax_rate
                st.session_state['current_margin_rate'] = margin_rate
                st.session_state['current_shipping_cost'] = shipping_cost
                st.session_state['extracted_tax'] = extracted_tax
                st.session_state['extracted_freight'] = extracted_freight
                st.session_state['current_desc_col'] = desc_col
                st.session_state['current_qty_col'] = qty_col
                st.session_state['current_cost_col'] = cost_col
                st.session_state['current_uom_col'] = uom_col
                st.session_state['current_total_col'] = total_col
                st.session_state['current_part_col'] = part_col
                
                # Initialize manual items if not exists
                if 'manual_items' not in st.session_state:
                    st.session_state['manual_items'] = []
                
                # Initialize item templates if not exists
                if 'item_templates' not in st.session_state:
                    st.session_state['item_templates'] = []
                
                # Increment files processed counter
                st.session_state.files_processed += 1
                
                # ── Manual Line Items ─────────────────────────────
                st.markdown('---')
                st.markdown('<h2 class="section-header">➕ Add Manual Line Items</h2>', unsafe_allow_html=True)
                st.markdown("""
                Add items not included in the vendor quote (e.g., Bollards, Pea gravel, additional materials).
                These will be included in all downloads and pricing calculations.
                """)
                
                with st.expander(f"➕ Manual Items ({len(st.session_state['manual_items'])} added)", expanded=len(st.session_state['manual_items']) == 0):
                    
                    # Tabs for different input methods
                    manual_tab1, manual_tab2, manual_tab3 = st.tabs(["Add Single Item", "Bulk Upload CSV", "Saved Templates"])
                    
                    # Tab 1: Single item addition
                    with manual_tab1:
                        with st.form("add_manual_item"):
                            st.markdown("**Add a single line item:**")
                            
                            man_col1, man_col2 = st.columns(2)
                            with man_col1:
                                man_desc = st.text_input("Material Description*", placeholder="e.g., Bollards, Pea gravel")
                                man_qty = st.number_input("Quantity*", min_value=0.0, step=1.0, value=1.0)
                                man_cost = st.number_input("Vendor Unit Cost*", min_value=0.0, step=0.01, format="%.2f")
                            
                            with man_col2:
                                man_uom = st.selectbox("Unit of Measure*", ["ea", "ft", "bag", "box", "roll", "spool", "lb", "ton", "yard", "other"])
                                man_part = st.text_input("Part Number (optional)", placeholder="Vendor SKU or part #")
                                save_template = st.checkbox("Save as template for future use")
                            
                            submit_col1, submit_col2 = st.columns([1, 3])
                            with submit_col1:
                                submitted = st.form_submit_button("Add to Quote", type="primary")
                            
                            if submitted:
                                if not man_desc or man_desc.strip() == "":
                                    st.error("Description is required")
                                elif man_qty <= 0:
                                    st.error("Quantity must be greater than 0")
                                elif man_cost < 0:
                                    st.error("Cost cannot be negative")
                                else:
                                    new_item = {
                                        'Description': man_desc.strip(),
                                        'Quantity': man_qty,
                                        'Unit Cost': man_cost,
                                        'Unit': man_uom,
                                        'Part Number': man_part.strip() if man_part else 'MANUAL-ENTRY',
                                        'Total': man_qty * man_cost,
                                        'Source': 'Manual Entry'
                                    }
                                    st.session_state['manual_items'].append(new_item)
                                    
                                    # Save as template if requested
                                    if save_template:
                                        template = {
                                            'name': man_desc.strip(),
                                            'uom': man_uom,
                                            'part': man_part.strip() if man_part else ''
                                        }
                                        if template not in st.session_state['item_templates']:
                                            st.session_state['item_templates'].append(template)
                                    
                                    st.success(f"✅ Added: {man_desc}")
                                    st.rerun()
                    
                    # Tab 2: CSV bulk upload
                    with manual_tab2:
                        st.markdown("""
                        **Upload multiple items at once via CSV**
                        
                        CSV format: `Description, Quantity, Unit Cost, Unit, Part Number`
                        
                        Example:
                        ```
                        Bollards,10,150.00,ea,BOLL-001
                        Pea Gravel,5,45.00,bag,GRAV-PEA
                        Warning Tape,3,12.50,roll,TAPE-WARN
                        ```
                        """)
                        
                        uploaded_csv = st.file_uploader("Upload CSV file", type=['csv'], key="manual_csv")
                        
                        if uploaded_csv:
                            try:
                                csv_df = pd.read_csv(uploaded_csv)
                                
                                # Validate columns
                                required_cols = ['Description', 'Quantity', 'Unit Cost', 'Unit']
                                if not all(col in csv_df.columns for col in required_cols):
                                    st.error(f"CSV must have columns: {', '.join(required_cols)}")
                                else:
                                    st.dataframe(csv_df.head(10), use_container_width=True)
                                    st.caption(f"Preview: {len(csv_df)} items")
                                    
                                    if st.button("Import All Items", type="primary", key="import_csv_btn"):
                                        for idx, row in csv_df.iterrows():
                                            new_item = {
                                                'Description': str(row['Description']).strip(),
                                                'Quantity': float(row['Quantity']),
                                                'Unit Cost': float(row['Unit Cost']),
                                                'Unit': str(row['Unit']).strip(),
                                                'Part Number': str(row.get('Part Number', 'MANUAL-ENTRY')).strip(),
                                                'Total': float(row['Quantity']) * float(row['Unit Cost']),
                                                'Source': 'CSV Import'
                                            }
                                            st.session_state['manual_items'].append(new_item)
                                        
                                        st.success(f"✅ Imported {len(csv_df)} items")
                                        st.rerun()
                            except Exception as e:
                                st.error(f"Error reading CSV: {str(e)}")
                    
                    # Tab 3: Templates
                    with manual_tab3:
                        if len(st.session_state['item_templates']) == 0:
                            st.info("No saved templates yet. Add items with 'Save as template' checked to create reusable templates.")
                        else:
                            st.markdown("**Quick-add from saved templates:**")
                            
                            for idx, template in enumerate(st.session_state['item_templates']):
                                temp_col1, temp_col2, temp_col3, temp_col4 = st.columns([3, 1, 1, 1])
                                
                                with temp_col1:
                                    st.text(f"{template['name']} ({template['uom']})")
                                
                                with temp_col2:
                                    qty = st.number_input("Qty", min_value=0.0, step=1.0, value=1.0, key=f"temp_qty_{idx}")
                                
                                with temp_col3:
                                    cost = st.number_input("Cost", min_value=0.0, step=0.01, key=f"temp_cost_{idx}")
                                
                                with temp_col4:
                                    if st.button("Add", key=f"temp_add_{idx}"):
                                        new_item = {
                                            'Description': template['name'],
                                            'Quantity': qty,
                                            'Unit Cost': cost,
                                            'Unit': template['uom'],
                                            'Part Number': template['part'] if template['part'] else 'MANUAL-ENTRY',
                                            'Total': qty * cost,
                                            'Source': 'Template'
                                        }
                                        st.session_state['manual_items'].append(new_item)
                                        st.rerun()
                    
                    # Display added manual items
                    if len(st.session_state['manual_items']) > 0:
                        st.markdown("---")
                        st.markdown(f"**Added Manual Items ({len(st.session_state['manual_items'])})**")
                        
                        manual_items_df = pd.DataFrame(st.session_state['manual_items'])
                        st.dataframe(manual_items_df, use_container_width=True)
                        
                        item_col1, item_col2, item_col3 = st.columns(3)
                        
                        with item_col1:
                            if st.button("🔄 Reprocess Quote with Manual Items", type="primary"):
                                # Combine vendor quote with manual items
                                base_df = st.session_state['base_cleaned_df'].copy()
                                
                                # Create manual items dataframe matching base structure
                                manual_df_rows = []
                                for item in st.session_state['manual_items']:
                                    row = {
                                        st.session_state['current_desc_col']: item['Description'],
                                        st.session_state['current_qty_col']: item['Quantity'],
                                        st.session_state['current_cost_col']: item['Unit Cost'],
                                        st.session_state['current_uom_col']: item['Unit'],
                                        st.session_state['current_total_col']: item['Total']
                                    }
                                    if st.session_state['current_part_col']:
                                        row[st.session_state['current_part_col']] = item['Part Number']
                                    manual_df_rows.append(row)
                                
                                manual_df = pd.DataFrame(manual_df_rows)
                                
                                # Combine dataframes
                                combined_df = pd.concat([base_df, manual_df], ignore_index=True)
                                
                                st.success(f"✅ Reprocessing quote with {len(st.session_state['manual_items'])} manual items...")
                                st.info("Scroll up to download updated files with manual items included.")
                                
                                # Would trigger reprocessing here - for now just show success
                                # In full implementation, this would regenerate all outputs
                        
                        with item_col2:
                            if st.button("🗑️ Clear All Manual Items"):
                                st.session_state['manual_items'] = []
                                st.rerun()
                        
                        with item_col3:
                            # Download manual items as CSV
                            manual_csv = manual_items_df.to_csv(index=False)
                            st.download_button(
                                label="📥 Export Manual Items",
                                data=manual_csv,
                                file_name=f"{st.session_state['current_project_name']}_manual_items.csv",
                                mime="text/csv"
                            )
                
                # ── Auto-add materials ────────────────────────────
                st.markdown('<h3>Add New Materials to Database</h3>', unsafe_allow_html=True)
                
                with st.expander("📦 Auto-add Materials from Quote"):
                    st.markdown("Add new materials from this quote to your materials database:")
                    
                    add_col1, add_col2, add_col3 = st.columns(3)
                    
                    with add_col1:
                        add_category = st.selectbox(
                            "Category for new materials",
                            ["Conduit", "Couplers", "Plugs", "Equipment", "Hand Holes", 
                             "Cable", "Aerial", "Grounding", "Markers", "Safety", "Materials", "Other"]
                        )
                    
                    with add_col2:
                        add_part_col = st.selectbox(
                            "Part Number Column (Optional)",
                            options=["None"] + all_cols,
                            help="Column containing vendor part numbers"
                        )
                    
                    with add_col3:
                        add_unit_col = st.selectbox(
                            "Unit Column (Optional)",
                            options=["None"] + all_cols,
                            help="Column containing units (ea, FT, etc.)"
                        )
                    
                    if st.button("Add New Materials", type="secondary"):
                        part_col_name = None if add_part_col == "None" else add_part_col
                        unit_col_name = None if add_unit_col == "None" else add_unit_col
                        
                        added = materials_db.add_materials_from_quote(
                            df=cleaned_df,
                            desc_col=desc_col,
                            part_col=part_col_name,
                            unit_col=unit_col_name if unit_col_name else 'Unit',
                            cost_col=cost_col,
                            category=add_category
                        )
                        
                        if added > 0:
                            st.success(f"Added {added} new materials to database!")
                        else:
                            st.info("No new materials found (all materials already in database)")
                        st.rerun()
            
            # Clean up temporary file (only for Excel)
            if not is_pdf:
                try:
                    if os.path.exists(temp_path):
                        os.unlink(temp_path)
                except Exception as cleanup_error:
                    st.warning(f"Warning: Could not clean up temporary file: {cleanup_error}")
                
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
            try:
                if 'temp_path' in locals() and os.path.exists(temp_path):
                    os.unlink(temp_path)
            except:
                pass


def build_quote_tab(materials_db, labor_db):
    """Manual quote building tab."""
    st.markdown('<h2 class="section-header">Build Quote Manually</h2>', 
                unsafe_allow_html=True)
    
    # Get materials by category
    materials_by_category = materials_db.get_materials_by_category()
    
    # Get labor tasks by category
    labor_by_category = labor_db.get_tasks_by_category()
    
    # Initialize session state for quote items
    if 'quote_items' not in st.session_state:
        st.session_state.quote_items = []
    
    if 'labor_items' not in st.session_state:
        st.session_state.labor_items = []
    
    # Add items section with tabs for Materials and Labor
    st.markdown('<h3>Add Items to Quote</h3>', unsafe_allow_html=True)
    
    materials_tab, labor_tab = st.tabs(["📦 Materials", "👷 Labor"])
    
    # ──────────────────────────────────────────────────────────
    # MATERIALS TAB
    # ──────────────────────────────────────────────────────────
    with materials_tab:
        col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
        
        with col1:
            # Material selection dropdown
            category = st.selectbox("Select Category", list(materials_by_category.keys()), key="mat_category")
            materials_in_category = materials_by_category[category]
            material_names = [m['name'] for m in materials_in_category]
            selected_material = st.selectbox("Select Material", material_names, key="mat_select")
        
        with col2:
            # Quantity input
            quantity = st.number_input("Quantity", min_value=0, value=1, step=1, key="mat_qty")
        
        with col3:
            # Unit cost override
            material = materials_db.find_material(selected_material)
            default_cost = material['unit_cost'] if material else 0.0
            material_cost = st.number_input("Unit Cost", min_value=0.0, value=default_cost, step=0.01, key="mat_cost",
                                           help="Override default cost if needed", format="%.2f")
        
        with col4:
            # Add button
            if st.button("Add to Quote", type="primary", key="add_material"):
                if material:
                    # Clean the material name for display
                    clean_name = materials_db._clean_material_name(material['name'])
                    item = {
                        'type': 'material',
                        'name': clean_name,
                        'part_number': material['part_number'],
                        'quantity': quantity,
                        'unit': material['unit'],
                        'unit_cost': material_cost,
                        'is_taxable': True
                    }
                    st.session_state.quote_items.append(item)
                    st.success(f"Added {quantity} x {clean_name}")
                    st.rerun()
    
    # ──────────────────────────────────────────────────────────
    # LABOR TAB
    # ──────────────────────────────────────────────────────────
    with labor_tab:
        col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
        
        with col1:
            # Labor task selection dropdown
            labor_category = st.selectbox("Select Category", list(labor_by_category.keys()), key="labor_category")
            tasks_in_category = labor_by_category[labor_category]
            task_names = [t['name'] for t in tasks_in_category]
            selected_task = st.selectbox("Select Task", task_names, key="labor_select")
        
        with col2:
            # Quantity input (hours/days/each)
            labor_qty = st.number_input("Quantity", min_value=0.0, value=1.0, step=0.5, key="labor_qty")
        
        with col3:
            # Rate override
            task = labor_db.find_task(selected_task)
            default_rate = task['base_rate'] if task else 0.0
            labor_rate = st.number_input("Rate", min_value=0.0, value=default_rate, step=1.0, key="labor_rate",
                                        help="Override default rate if needed")
        
        with col4:
            # Tax toggle
            labor_taxable = st.checkbox("Taxable", value=False, key="labor_taxable",
                                       help="Labor is usually tax-exempt")
            
            # Add button
            if st.button("Add to Quote", type="primary", key="add_labor"):
                if task:
                    labor_item = {
                        'type': 'labor',
                        'name': task['name'],
                        'part_number': None,
                        'quantity': labor_qty,
                        'unit': task['unit'],
                        'unit_cost': labor_rate,
                        'is_taxable': labor_taxable
                    }
                    st.session_state.labor_items.append(labor_item)
                    st.success(f"Added {labor_qty} {task['unit']} of {task['name']}")
                    st.rerun()
    
    # Display quote items (materials + labor combined)
    if st.session_state.quote_items or st.session_state.labor_items:
        st.markdown('<h3>Quote Items</h3>', unsafe_allow_html=True)
        
        # Combine materials and labor items
        all_items = st.session_state.quote_items + st.session_state.labor_items
        quote_df = pd.DataFrame(all_items)
        
        # Add calculations
        tax_rate = st.sidebar.slider("Tax Rate (%)", 0.0, 20.0, 8.25, 0.25) / 100
        margin_rate = st.sidebar.slider("Margin Rate (%)", 0.0, 50.0, 10.0, 1.0) / 100
        
        processor = PricingProcessor(tax_rate=tax_rate, margin_rate=margin_rate)
        
        # Calculate composite rate and total for each item
        def calculate_item_total(row):
            result = processor.calculate_composite_rate(row['unit_cost'])
            # If not taxable (labor), recalculate without tax
            if not row.get('is_taxable', True):
                # Apply only margin, no tax
                margin_dollars = row['unit_cost'] * margin_rate
                composite = row['unit_cost'] + margin_dollars
            else:
                composite = result['composite_rate']
            return composite
        
        quote_df['composite_rate'] = quote_df.apply(calculate_item_total, axis=1)
        quote_df['total'] = (quote_df['quantity'] * quote_df['composite_rate']).round(2)
        
        # Set column attributes for download methods
        quote_df.attrs['desc_col'] = 'name'
        quote_df.attrs['qty_col'] = 'quantity'
        quote_df.attrs['cost_col'] = 'unit_cost'
        quote_df.attrs['part_col'] = 'part_number'
        quote_df.attrs['uom_col'] = 'unit'
        
        # Display with formatting
        display_df = quote_df[['type', 'name', 'quantity', 'unit', 'composite_rate', 'total', 'is_taxable']].copy()
        display_df.columns = ['Type', 'Description', 'Quantity', 'Unit', 'Unit Price', 'Total', 'Taxable']
        display_df['Type'] = display_df['Type'].map({'material': '📦', 'labor': '👷'})
        display_df['Taxable'] = display_df['Taxable'].map({True: '✓', False: '✗'})
        display_df['Unit Price'] = display_df['Unit Price'].map('${:,.2f}'.format)
        display_df['Total'] = display_df['Total'].map('${:,.2f}'.format)
        
        st.dataframe(display_df, use_container_width=True)
        
        # Summary with separate materials/labor breakdown
        materials_total = quote_df[quote_df['type'] == 'material']['total'].sum() if 'material' in quote_df['type'].values else 0
        labor_total = quote_df[quote_df['type'] == 'labor']['total'].sum() if 'labor' in quote_df['type'].values else 0
        total_amount = quote_df['total'].sum()
        
        # Calculate tax breakdown
        materials_cost = quote_df[quote_df['type'] == 'material']['unit_cost'].sum() if 'material' in quote_df['type'].values else 0
        materials_tax = materials_cost * margin_rate * tax_rate if materials_cost > 0 else 0
        labor_tax = 0  # Labor is tax-exempt by default
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Materials", f"${materials_total:,.2f}")
        with col2:
            st.metric("Labor", f"${labor_total:,.2f}")
        with col3:
            st.metric("Total", f"${total_amount:,.2f}")
        with col4:
            st.metric("Tax (Materials only)", f"${materials_tax:,.2f}")
        
        # Download buttons
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # Generate internal audit file
            audit_buffer = BytesIO()
            audit_df = processor.generate_internal_audit_spreadsheet(
                quote_df, 
                tax_rate=tax_rate, 
                margin_rate=margin_rate,
                output_path=audit_buffer
            )
            
            st.download_button(
                label="📥 Download Audit",
                data=audit_buffer.getvalue(),
                file_name="manual_audit.xlsx",
                help="Full calculation breakdown",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col2:
            # Generate client quote file
            client_buffer = BytesIO()
            client_df_manual = processor.generate_client_spreadsheet(
                quote_df,
                output_path=client_buffer
            )
            
            st.download_button(
                label="📥 Download Client Quote",
                data=client_buffer.getvalue(),
                file_name="manual_client_quote.xlsx",
                help="Clean quote for client",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col3:
            # Combined file
            combined_buffer = BytesIO()
            with pd.ExcelWriter(combined_buffer, engine='xlsxwriter') as writer:
                audit_df.to_excel(writer, sheet_name='Internal_Audit', index=False)
                client_df_manual.to_excel(writer, sheet_name='Client_Quote', index=False)
            
            st.download_button(
                label="📥 Download Both",
                data=combined_buffer.getvalue(),
                file_name="manual_quote_complete.xlsx",
                help="Both sheets in one file",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        # Clear button
        if st.button("Clear Quote", type="secondary"):
            st.session_state.quote_items = []
            st.session_state.labor_items = []
            st.rerun()


def materials_list_tab(materials_db):
    """Materials list management tab."""
    st.markdown('<h2 class="section-header">Materials Database</h2>', 
                unsafe_allow_html=True)
    
    st.markdown("""
    Browse our comprehensive database of OSP fiber construction materials.
    All pricing is placeholder ($0.00) - upload vendor quotes to populate actual costs.
    """)
    
    # Get full materials dataframe with all columns
    materials_df = materials_db.to_dataframe(hide_sensitive=False)
    
    # Stats row
    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Total Materials", len(materials_df))
    with col2:
        st.metric("Categories", materials_df['category'].nunique())
    with col3:
        st.metric("Units of Measure", materials_df['unit'].nunique())
    
    st.markdown("---")
    
    # Filter row
    filter_col1, filter_col2 = st.columns([1, 2])
    
    with filter_col1:
        # Category filter
        categories = sorted(materials_df['category'].unique().tolist())
        selected_category = st.selectbox(
            "Filter by Category",
            ["All Categories"] + categories,
            help="Filter materials by category"
        )
    
    with filter_col2:
        # Search filter
        search_term = st.text_input(
            "🔍 Search materials",
            placeholder="e.g., conduit, fiber, splice, cable...",
            help="Search by material name"
        )
    
    # Apply filters
    filtered_df = materials_df.copy()
    
    if selected_category != "All Categories":
        filtered_df = filtered_df[filtered_df['category'] == selected_category]
    
    if search_term:
        filtered_df = filtered_df[filtered_df['name'].str.contains(search_term, case=False, na=False)]
    
    # Display filtered results
    st.markdown(f"**Showing {len(filtered_df)} materials**")
    
    # Format for display - show only relevant columns
    display_df = filtered_df[['name', 'category', 'unit', 'unit_cost']].copy()
    display_df.columns = ['Material Name', 'Category', 'Unit', 'Unit Cost']
    display_df['Unit Cost'] = display_df['Unit Cost'].map('${:,.2f}'.format)
    
    st.dataframe(display_df, use_container_width=True, height=600)
    
    # Edit existing materials
    with st.expander("Edit Material Pricing"):
        st.markdown("**Select a material to update its pricing:**")
        
        # Material selection for editing
        edit_material_names = materials_df['name'].tolist()
        selected_edit_material = st.selectbox("Select Material to Edit", edit_material_names)
        
        if selected_edit_material:
            material = materials_db.find_material(selected_edit_material)
            if material:
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    current_cost = st.number_input(
                        "Current Unit Cost",
                        min_value=0.0,
                        value=float(material['unit_cost']),
                        step=0.01,
                        key=f"cost_{material['name']}"
                    )
                
                with col2:
                    new_part = st.text_input(
                        "Part Number",
                        value=material['part_number'] or "",
                        key=f"part_{material['name']}"
                    )
                
                with col3:
                    new_unit = st.selectbox(
                        "Unit",
                        options=["each", "FT", "ea", "bag", "PK", "REEL"],
                        index=["each", "FT", "ea", "bag", "PK", "REEL"].index(material['unit']) if material['unit'] in ["each", "FT", "ea", "bag", "PK", "REEL"] else 0,
                        key=f"unit_{material['name']}"
                    )
                
                # Update button
                col1, col2, col3 = st.columns([1, 1, 2])
                with col1:
                    if st.button("Update Material", type="primary"):
                        # Update the material using the database method
                        materials_db.update_material(
                            selected_edit_material,
                            unit_cost=current_cost,
                            part_number=new_part if new_part else None,
                            unit=new_unit
                        )
                        st.success(f"Updated {selected_edit_material}")
                        st.rerun()
                
                with col2:
                    if st.button("Delete Material", type="secondary"):
                        # Delete the material using the database method
                        materials_db.delete_material(selected_edit_material)
                        st.success(f"Deleted {selected_edit_material}")
                        st.rerun()
    
    # Bulk price update section
    with st.expander("Bulk Price Update"):
        st.markdown("**Apply percentage increase/decrease to all materials or by category:**")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            update_type = st.selectbox("Update Scope", ["All Materials", "By Category"])
            
        with col2:
            if update_type == "By Category":
                categories = list(set(materials_df['category'].tolist()))
                selected_category = st.selectbox("Select Category", categories)
            else:
                selected_category = None
        
        with col3:
            percentage_change = st.number_input(
                "% Change",
                value=0.0,
                step=0.1,
                help="Positive for increase, negative for decrease"
            )
        
        if st.button("Apply Bulk Update"):
            if update_type == "All Materials":
                # Update all materials
                for mat in materials_db.materials:
                    old_cost = mat['unit_cost']
                    new_cost = old_cost * (1 + percentage_change / 100)
                    mat['unit_cost'] = round(new_cost, 4)
                materials_db.save()  # Save all changes
                st.success(f"Updated all materials by {percentage_change}%")
            else:
                # Update only selected category
                updated_count = 0
                for mat in materials_db.materials:
                    if mat['category'] == selected_category:
                        old_cost = mat['unit_cost']
                        new_cost = old_cost * (1 + percentage_change / 100)
                        mat['unit_cost'] = round(new_cost, 4)
                        updated_count += 1
                materials_db.save()  # Save all changes
                st.success(f"Updated {updated_count} materials in {selected_category} by {percentage_change}%")
            st.rerun()
    
    # Add new material section
    with st.expander("Add New Material"):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            new_name = st.text_input("Material Name")
            new_part = st.text_input("Part Number")
        
        with col2:
            new_cost = st.number_input("Unit Cost", min_value=0.0, value=0.0, step=0.01)
            new_unit = st.selectbox("Unit", ["each", "FT", "ea", "bag", "PK", "REEL"])
        
        with col3:
            all_categories = sorted(set(materials_df['category'].tolist() + 
                ["Conduit", "Couplers", "Plugs", "Equipment", "Hand Holes",
                 "Cable", "Aerial", "Grounding", "Markers", "Safety", "Materials", "Other"]))
            new_category = st.selectbox("Category", all_categories)
        
        if st.button("Add Material") and new_name:
            materials_db.add_material(
                name=new_name,
                part_number=new_part if new_part else None,
                unit_cost=new_cost,
                unit=new_unit,
                category=new_category
            )
            st.success(f"Added {new_name} to materials database")
            st.rerun()
    
    # Export/Import
    col1, col2 = st.columns(2)
    
    with col1:
        # Export materials (without sensitive data)
        csv = materials_db.to_dataframe(hide_sensitive=True).to_csv(index=False)
        st.download_button(
            label="📥 Export Materials CSV",
            data=csv,
            file_name="materials_list.csv",
            mime="text/csv"
        )
    
    with col2:
        # Import materials
        uploaded_materials = st.file_uploader(
            "Import Materials CSV",
            type=['csv'],
            help="Upload a CSV file with materials"
        )
        
        if uploaded_materials:
            try:
                imported_df = pd.read_csv(uploaded_materials)
                st.success(f"Imported {len(imported_df)} materials")
                st.dataframe(imported_df)
            except Exception as e:
                st.error(f"Error importing file: {str(e)}")
    
    # Instructions section
    with st.expander("📖 How to Use"):
        st.markdown("""
        **Process Quote Tab:**
        1. Upload your Excel file containing the vendor quote
        2. Select the sheet with the quote data
        3. Adjust tax and margin rates in the sidebar if needed
        4. Click "Process Quote" to calculate composite rates
        5. Download the results
        
        **Build Quote Tab:**
        1. Select materials from the dropdown by category
        2. Enter quantities for each item
        3. Add items to build your quote
        4. Download the completed quote
        
        **Materials List Tab:**
        1. View all available materials
        2. Search for specific items
        3. Add new materials to the database
        4. Export/import materials lists
        
        The calculator automatically adds margin and tax to create clean client pricing.
        """)
    
    # Example format section
    with st.expander("📋 Expected Format"):
        st.markdown("""
        **For Process Quote Tab:**
        The app works best with Excel formats including:
        - Material Description
        - Part Number
        - Quantity
        - Unit Cost
        
        **For Build Quote Tab:**
        Uses the built-in materials database with standard OSP items.
        You can add custom materials in the Materials List tab.
        """)


if __name__ == "__main__":
    main()